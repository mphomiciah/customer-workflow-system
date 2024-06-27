from flask import Flask, render_template, request, redirect, url_for, flash
import os
import uuid
import sqlite3
from openpyxl import load_workbook
import matplotlib.pyplot as plt

app = Flask(__name__)
app.secret_key = 'supersecretkey'

UPLOAD_FOLDER = 'uploads'
STATIC_FOLDER = 'static'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['STATIC_FOLDER'] = STATIC_FOLDER
DATABASE = 'database.db'

# function to establish a connection to SQLite database
def get_db():
    conn = sqlite3.connect(DATABASE)
    return conn

# route for the index page
@app.route('/')
def index():
    return render_template('index.html')

# handling file upload and form submission
@app.route('/upload', methods=['POST'])
def upload_file():
    if 'fileUpload' not in request.files:
        flash('No file part')
        return redirect(request.url)
    file = request.files['fileUpload']
    if file.filename == '':
        flash('No selected file')
        return redirect(request.url)

    # to only allow excel files (.xls and .xlsx)
    if file and file.filename.endswith(('.xls', '.xlsx')):
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        #get form data
        fname = request.form['fname']
        lname = request.form['lname']
        birthday = request.form['dateOfBirth']
        
        # Load workbook and active sheet from uploaded Excel file
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active

        #connect database
        conn = get_db()
        cursor = conn.cursor()

        # Create employee table if it not exist
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS employee (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                first_name TEXT,
                last_name TEXT,
                birthday TEXT,
                month TEXT,
                income REAL,
                expense REAL
            )
        ''')

        # loop through rows in excel sheet starting from row 2 (skipping headers)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[2]:  # Ensure there are no empty cells
                month = row[0]
                
                # handle currency formatting before converting to float
                income = float(str(row[1]).replace('R', '').replace(',', '').strip())
                expense = float(str(row[2]).replace('R', '').replace(',', '').strip())
                
                 # insert data into database table
                cursor.execute('''
                    INSERT INTO employee (first_name, last_name, birthday, month, income, expense)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (fname, lname, birthday, month, income, expense))

        # commit changes to database and close connection
        conn.commit()
        conn.close()
        
        flash('File successfully uploaded and processed')

        # generate a unique filename for the graph
        graph_filename = f'graph_{uuid.uuid4().hex}.png'
        graph_path = os.path.join(app.config['STATIC_FOLDER'], 'graphs', graph_filename)

        # call function to generate graph based on data in database
        generate_graph(graph_path)

        return redirect(url_for('graph', graph_filename=graph_filename))  # Redirect to the graph page with the graph filename
    else:
        flash('Invalid file type')
        return redirect(request.url)

@app.route('/graph')
def graph():
    graph_filename = request.args.get('graph_filename')
    if graph_filename:
        graph_path = url_for('static', filename=f'graphs/{graph_filename}')
    else:
        graph_path = url_for('static', filename='placeholder.png')  # Placeholder if no graph generated yet
    return render_template('graph.html', graph_path=graph_path)

def generate_graph(graph_path):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT month, SUM(income), SUM(expense) FROM employee GROUP BY month')
    data = cursor.fetchall()
    conn.close()
    
    # order our months from Jan to Dec
    months_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    chart_data = {'labels': [], 'income': [], 'expense': []}
    monthly_data = {month: (0, 0) for month in months_order}  # Initialize with zero values

    for row in data:
        monthly_data[row[0]] = (row[1], row[2])

    for month in months_order:
        chart_data['labels'].append(month)
        chart_data['income'].append(monthly_data[month][0])
        chart_data['expense'].append(monthly_data[month][1])

    # color for income and expense bars
    income_color = 'lightblue'
    expense_color = 'lightpink'

    # Generate the bar chart using matplotlib
    bar_width = 0.35
    index = range(len(chart_data['labels']))

    fig, ax = plt.subplots(figsize=(10, 6))
    bar1 = ax.bar(index, chart_data['income'], bar_width, label='Income', color=income_color)
    bar2 = ax.bar([i + bar_width for i in index], chart_data['expense'], bar_width, label='Expenses', color=expense_color)

    ax.set_xlabel('Months')
    ax.set_ylabel('Amount')
    ax.set_title('Income and Expenses Over Last 12 Months')
    ax.set_xticks([i + bar_width / 2 for i in index])
    ax.set_xticklabels(chart_data['labels'])
    ax.legend()

    plt.tight_layout()

    # Save the plot to the specified graph path
    plt.savefig(graph_path)
    plt.close()

if __name__ == '__main__':
    # Initialize database if it doesn't exist
    if not os.path.exists(DATABASE):
        conn = sqlite3.connect(DATABASE)
        conn.close()
    
    # Create upload and static folders if they don't exist
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    if not os.path.exists(os.path.join(STATIC_FOLDER, 'graphs')):
        os.makedirs(os.path.join(STATIC_FOLDER, 'graphs'))
    app.run(debug=True)
